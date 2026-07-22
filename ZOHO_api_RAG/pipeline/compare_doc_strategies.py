"""
compare_doc_strategies.py

Compares chunk-text strategies for the Site24x7 RAG corpus:

  - query_derived   : current approach (rag_eval.build_chunk_text) - method/endpoint,
                       humanized path words, sheet>subFeature category, N example
                       queries pulled from the enrichment pool. No external doc needed.
  - doc_description : method/endpoint + the description-file text for that
                       (endpoint, method), when available. Falls back to
                       query_derived text for the ~78% of pairs the doc file
                       doesn't cover (it has no "Reports" sheet at all).
  - concat           : query_derived text + the description-file text appended,
                       when available. Falls back to query_derived alone otherwise.

Only ~21.5% of (endpoint, method) pairs in the dataset have a doc-file entry
(the doc file is missing a "Reports" sheet entirely, which is ~79% of the
dataset). Because of that, doc_description / concat scores on the FULL eval
set are mostly measuring "did we fall back correctly", not "is the doc text
better". To isolate doc-text quality on its own, this script also runs a
DOC-COVERED-ONLY eval: same three strategies, but the eval query sample is
drawn only from queries whose true (endpoint, method) is one the doc file
actually covers, and doc_description uses no fallback there.

Overfitting control: rather than the single fixed seed=42 / n_eval=100 sample
rag_eval.py normally uses, this script repeats the eval across several
different seeds (same n_eval each time) and reports mean +/- std of
Recall@1/@10 per strategy, plus the Reports vs non-Reports split, so a
strategy that only looks better because it "won" one lucky sample is visible
as high-variance rather than a real win.

This script does NOT modify rag_eval.py, diagnose_retrieval.py, or
populate_api_docs.py - it imports the corpus/eval helpers and embedding
helpers from rag_eval.py and reuses them as-is.

Usage:
    python compare_doc_strategies.py site24x7_Dataset.csv site24x7_Admin_API.xlsx --mock
    python compare_doc_strategies.py site24x7_Dataset.csv site24x7_Admin_API.xlsx \
        --base-url http://20.235.183.15:443/openai/v1 --api-key YOUR_KEY \
        --seeds 1 2 3 4 5 6 7 8 --n-eval 100
"""

import argparse
import sys
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

from rag_eval import (
    build_chunk_text,
    build_corpus_and_eval,
    compute_recall,
    embed_texts,
    get_embeddings_mock,
    humanize_path,
)


# ---------- description-file loading ----------

def load_doc_descriptions(xlsx_path: Path) -> dict:
    """Parse every sheet of the Admin API description workbook into a dict
    keyed by (endpoint, method) -> best available description text.

    Each sheet has a title row, a free-text description row, a blank row,
    then a header row ('Sub-Feature', 'Endpoint Path', 'HTTP Method',
    'Status Code', "What It's For (Sub-Feature)", ...), then data rows - one
    row per (endpoint, method, status_code). A given (endpoint, method) can
    appear multiple times (e.g. 200 vs 401 rows) with different description
    text; we keep the LONGEST non-empty description seen for that pair, on
    the assumption the richer one is more useful as retrieval-chunk text.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    best: dict = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        header_idx = None
        for i, r in enumerate(rows[:6]):
            if r and r[0] == "Sub-Feature":
                header_idx = i
                break
        if header_idx is None:
            print(f"  [doc-file] WARNING: no header row found in sheet '{sheet_name}', skipping")
            continue

        for r in rows[header_idx + 1:]:
            if r is None or r[0] is None or r[1] is None:
                continue
            endpoint, method, description = r[1], r[2], r[4]
            if not endpoint or not method:
                continue
            key = (str(endpoint).strip(), str(method).strip())
            desc = str(description).strip() if description else ""
            if not desc:
                continue
            if key not in best or len(desc) > len(best[key]):
                best[key] = desc

    return best


def load_extra_descriptions(csv_path: Path) -> dict:
    """Load a supplementary descriptions CSV (e.g. LLM-synthesized Reports-sheet
    descriptions from synthesize_reports_descriptions.py) with columns
    endpoint, method, description (extra columns are ignored). Returns a dict
    keyed the same way as load_doc_descriptions() so it can be merged in."""
    extra_df = pd.read_csv(csv_path)
    required = {"endpoint", "method", "description"}
    missing = required - set(extra_df.columns)
    if missing:
        sys.exit(f"--extra-descriptions file is missing required column(s): {missing}")
    out = {}
    for _, r in extra_df.iterrows():
        desc = str(r["description"]).strip() if pd.notna(r["description"]) else ""
        if not desc:
            continue
        out[(str(r["endpoint"]).strip(), str(r["method"]).strip())] = desc
    return out


# ---------- corpus text variants ----------

def build_text_query_derived(method, endpoint, sheet, sub_feature, examples, doc_desc):
    return build_chunk_text(method, endpoint, sheet, sub_feature, examples)


def build_text_doc_description(method, endpoint, sheet, sub_feature, examples, doc_desc):
    if not doc_desc:
        # fallback: doc file has no entry for this pair (very common - see docstring)
        return build_chunk_text(method, endpoint, sheet, sub_feature, examples)
    parts = [f"{method} {endpoint}", doc_desc]
    return "\n".join(parts)


def build_text_concat(method, endpoint, sheet, sub_feature, examples, doc_desc):
    base = build_chunk_text(method, endpoint, sheet, sub_feature, examples)
    if not doc_desc:
        return base
    return base + "\n" + doc_desc


STRATEGIES = {
    "query_derived": build_text_query_derived,
    "doc_description": build_text_doc_description,
    "concat": build_text_concat,
}


def make_variant_corpus(base_corpus_df: pd.DataFrame, df: pd.DataFrame, doc_map: dict,
                         n_examples: int, enrichment_pool: pd.DataFrame, strategy_fn) -> pd.DataFrame:
    """Rebuild the 'text' column of a corpus_df (same rows/keys/order as
    base_corpus_df, produced by rag_eval.build_corpus_and_eval) under a
    different text-building strategy. Sheet/subFeature/examples are pulled
    the same way build_corpus_and_eval does, so the only thing that differs
    across strategies is the text-building function."""
    rows = []
    for _, base_row in base_corpus_df.iterrows():
        endpoint, method = base_row["endpoint"], base_row["method"]
        key = base_row["key"]
        group = df[(df["endpoint"] == endpoint) & (df["method"] == method)]
        sheet = group["sheet"].iloc[0]
        sub_feature = group["subFeature"].iloc[0]
        pool = enrichment_pool[enrichment_pool["key"] == key]["query"].tolist()
        examples = pool[:n_examples]
        doc_desc = doc_map.get((endpoint, method))
        text = strategy_fn(method, endpoint, sheet, sub_feature, examples, doc_desc)
        rows.append({"endpoint": endpoint, "method": method, "key": key, "text": text})
    return pd.DataFrame(rows)


# ---------- embedding helper that dispatches mock vs real ----------

def embed(texts, client, args, cache_dir):
    if args.mock:
        return get_embeddings_mock(texts)
    return embed_texts(client, args.embed_model, texts, cache_dir, not args.no_cache,
                        args.embed_batch_size, request_timeout=args.timeout)


# ---------- stratified recall (Reports vs non-Reports) ----------

def stratified_recall(per_query_df: pd.DataFrame, eval_df: pd.DataFrame, sheet_of_key: dict, ks=(1, 10)):
    """per_query_df comes from rag_eval.compute_recall (has 'query', 'rank_of_correct').
    eval_df has 'query' and 'valid_keys'. We tag each eval query as Reports/non-Reports
    based on whether ANY of its valid_keys belongs to the Reports sheet, then report
    Recall@k separately for each group."""
    query_to_valid = dict(zip(eval_df["query"], eval_df["valid_keys"]))
    out = {}
    for group_name, predicate in [
        ("reports", lambda keys: any(sheet_of_key.get(k) == "Reports" for k in keys)),
        ("non_reports", lambda keys: not any(sheet_of_key.get(k) == "Reports" for k in keys)),
    ]:
        mask = per_query_df["query"].map(lambda q: predicate(query_to_valid.get(q, [])))
        subset = per_query_df[mask]
        n = len(subset)
        res = {}
        for k in ks:
            if n == 0:
                res[f"recall@{k}"] = float("nan")
            else:
                hits = (subset["rank_of_correct"].apply(lambda r: r != f">{max(ks)}" and int(r) <= k
                                                          if isinstance(r, str) else r <= k)).sum()
                res[f"recall@{k}"] = hits / n
        res["n"] = n
        out[group_name] = res
    return out


# ---------- doc-covered-only eval set ----------

def build_doc_covered_eval(df: pd.DataFrame, doc_keys: set, n_eval: int, seed: int):
    """Like rag_eval.build_corpus_and_eval's eval-set construction, but
    restricted to queries whose (single) true answer is a doc-covered pair.
    Keeps it simple: only queries with exactly one valid key are eligible,
    and that key must be in doc_keys, so doc_description never needs to
    fall back for this eval set."""
    true_df = df[df["markedCorrect"] == True].copy()
    true_df["key"] = list(zip(true_df["endpoint"], true_df["method"]))

    query_groups = (
        true_df.groupby("query")["key"]
        .apply(lambda keys: sorted(set(keys)))
        .reset_index()
        .rename(columns={"key": "valid_keys"})
    )
    eligible = query_groups[
        query_groups["valid_keys"].apply(lambda ks: len(ks) == 1 and ks[0] in doc_keys)
    ]
    n = min(n_eval, len(eligible))
    eval_df = eligible.sample(n=n, random_state=seed).reset_index(drop=True)
    return eval_df, len(eligible)


def main():
    parser = argparse.ArgumentParser(description="Compare corpus-text strategies across repeated eval samples")
    parser.add_argument("csv_path")
    parser.add_argument("xlsx_path")
    parser.add_argument("--n-eval", type=int, default=100)
    parser.add_argument("--n-examples", type=int, default=2)
    parser.add_argument("--seeds", type=int, nargs="+", default=[1, 2, 3, 4, 5, 6, 7, 8],
                         help="one eval run per seed; mean/std reported across them")
    parser.add_argument("--base-url", default=None)
    parser.add_argument("--api-key", default=None)
    parser.add_argument("--embed-model", default="azure:primary/s247-textembedding-3l")
    parser.add_argument("--mock", action="store_true")
    parser.add_argument("--cache-dir", default=None)
    parser.add_argument("--no-cache", action="store_true")
    parser.add_argument("--embed-batch-size", type=int, default=20)
    parser.add_argument("--timeout", type=float, default=60.0)
    parser.add_argument("--skip-doc-covered-only", action="store_true",
                         help="skip the second (doc-covered-only) eval pass")
    parser.add_argument("--extra-descriptions", default=None,
                         help="optional CSV (endpoint, method, description columns) to merge into the "
                              "doc_map alongside the xlsx - e.g. reports_synthetic_descriptions.csv from "
                              "synthesize_reports_descriptions.py, to extend coverage into sheets the xlsx "
                              "doesn't cover. Xlsx entries win on any key collision.")
    args = parser.parse_args()

    csv_path = Path(args.csv_path)
    xlsx_path = Path(args.xlsx_path)
    if not csv_path.exists():
        sys.exit(f"File not found: {csv_path}")
    if not xlsx_path.exists():
        sys.exit(f"File not found: {xlsx_path}")

    df = pd.read_csv(csv_path)
    if df["markedCorrect"].dtype == object:
        df["markedCorrect"] = df["markedCorrect"].astype(str).str.strip().str.lower().map(
            {"true": True, "false": False, "1": True, "0": False}
        )
    df["markedCorrect"] = df["markedCorrect"].astype(bool)

    print("Loading description file...")
    doc_map = load_doc_descriptions(xlsx_path)
    doc_keys = set(doc_map.keys())
    all_keys = set(zip(df["endpoint"], df["method"]))
    print(f"  doc file covers {len(doc_keys & all_keys)}/{len(all_keys)} unique (endpoint, method) pairs "
          f"({len(doc_keys & all_keys) / len(all_keys) * 100:.1f}%)")

    if args.extra_descriptions:
        extra_path = Path(args.extra_descriptions)
        if not extra_path.exists():
            sys.exit(f"File not found: {extra_path}")
        extra_map = load_extra_descriptions(extra_path)
        new_keys = set(extra_map.keys()) - doc_keys  # xlsx wins on collision
        for k in new_keys:
            doc_map[k] = extra_map[k]
        doc_keys = set(doc_map.keys())
        collisions = set(extra_map.keys()) & (doc_keys - new_keys)
        print(f"  + extra descriptions file adds {len(new_keys)} new pairs "
              f"({len(collisions)} keys already had an xlsx entry and were left as-is)")
        print(f"  combined doc coverage: {len(doc_keys & all_keys)}/{len(all_keys)} unique pairs "
              f"({len(doc_keys & all_keys) / len(all_keys) * 100:.1f}%)")
    print()

    sheet_of_key = {}
    for (ep, m), g in df.groupby(["endpoint", "method"]):
        sheet_of_key[(ep, m)] = g["sheet"].iloc[0]  # same arbitrary-first-row convention as build_corpus_and_eval

    cache_dir = Path(args.cache_dir) if args.cache_dir else csv_path.parent / ".rag_cache"

    client = None
    if not args.mock:
        if not args.base_url or not args.api_key:
            sys.exit("Provide --base-url and --api-key, or use --mock for an offline test.")
        from openai import OpenAI
        client = OpenAI(base_url=args.base_url, api_key=args.api_key, timeout=args.timeout)

    # ============ PASS 1: full eval set, repeated across seeds ============
    print("=" * 70)
    print("PASS 1: full eval set (all strategies fall back to query_derived")
    print("        text wherever the doc file has no entry)")
    print("=" * 70)

    full_rows = []
    full_per_query = []

    for seed in args.seeds:
        print(f"\n--- seed {seed} ---")
        base_corpus_df, eval_df = build_corpus_and_eval(df, args.n_eval, args.n_examples, seed)

        true_df = df[df["markedCorrect"] == True].copy()
        true_df["key"] = list(zip(true_df["endpoint"], true_df["method"]))
        eval_queries = set(eval_df["query"])
        enrichment_pool = true_df[~true_df["query"].isin(eval_queries)]

        query_vecs = embed(eval_df["query"].tolist(), client, args, cache_dir)

        for strat_name, strat_fn in STRATEGIES.items():
            variant_corpus_df = make_variant_corpus(base_corpus_df, df, doc_map, args.n_examples,
                                                      enrichment_pool, strat_fn)
            corpus_vecs = embed(variant_corpus_df["text"].tolist(), client, args, cache_dir)

            per_query_df, recall = compute_recall(variant_corpus_df, eval_df, corpus_vecs, query_vecs,
                                                    ks=(1, 10), variant=f"{strat_name}_seed{seed}")
            per_query_df["strategy"] = strat_name
            per_query_df["seed"] = seed
            full_per_query.append(per_query_df)

            strat = stratified_recall(per_query_df, eval_df, sheet_of_key, ks=(1, 10))
            full_rows.append({
                "strategy": strat_name, "seed": seed,
                "recall@1": recall[1], "recall@10": recall[10],
                "n_eval": len(eval_df),
                "reports_recall@1": strat["reports"]["recall@1"], "reports_n": strat["reports"]["n"],
                "non_reports_recall@1": strat["non_reports"]["recall@1"], "non_reports_n": strat["non_reports"]["n"],
            })

    full_results_df = pd.DataFrame(full_rows)
    out_dir = csv_path.parent
    full_results_df.to_csv(out_dir / "compare_doc_strategies_full_eval_per_seed.csv", index=False)
    pd.concat(full_per_query, ignore_index=True).to_csv(
        out_dir / "compare_doc_strategies_full_eval_per_query.csv", index=False)

    print("\n" + "=" * 70)
    print("PASS 1 SUMMARY (mean +/- std across seeds)")
    print("=" * 70)
    summary1 = full_results_df.groupby("strategy").agg(
        recall_1_mean=("recall@1", "mean"), recall_1_std=("recall@1", "std"),
        recall_10_mean=("recall@10", "mean"), recall_10_std=("recall@10", "std"),
        reports_recall_1_mean=("reports_recall@1", "mean"), reports_recall_1_std=("reports_recall@1", "std"),
        non_reports_recall_1_mean=("non_reports_recall@1", "mean"), non_reports_recall_1_std=("non_reports_recall@1", "std"),
    ).round(4)
    print(summary1.to_string())
    summary1.to_csv(out_dir / "compare_doc_strategies_full_eval_summary.csv")

    if args.skip_doc_covered_only:
        print(f"\nSaved -> {out_dir / 'compare_doc_strategies_full_eval_summary.csv'}")
        print("Skipping doc-covered-only pass (--skip-doc-covered-only set).")
        return

    # ============ PASS 2: doc-covered-only eval set ============
    print("\n" + "=" * 70)
    print("PASS 2: doc-covered-only eval set (isolates doc-text QUALITY from")
    print("        doc-text COVERAGE - doc_description never falls back here)")
    print("=" * 70)

    covered_rows = []
    covered_per_query = []
    n_eligible_reported = False

    for seed in args.seeds:
        eval_df, n_eligible = build_doc_covered_eval(df, doc_keys & all_keys, args.n_eval, seed)
        if not n_eligible_reported:
            print(f"  {n_eligible} queries in the dataset have a doc-covered single true answer "
                  f"(sampling {len(eval_df)} per seed)\n")
            n_eligible_reported = True
        print(f"--- seed {seed} ---")

        # corpus is still the FULL corpus (all 2298 pairs) - we're only restricting
        # which queries we evaluate, not what the model can retrieve
        base_corpus_df, _ = build_corpus_and_eval(df, args.n_eval, args.n_examples, seed)
        true_df = df[df["markedCorrect"] == True].copy()
        true_df["key"] = list(zip(true_df["endpoint"], true_df["method"]))
        eval_queries = set(eval_df["query"])
        enrichment_pool = true_df[~true_df["query"].isin(eval_queries)]

        query_vecs = embed(eval_df["query"].tolist(), client, args, cache_dir)

        for strat_name, strat_fn in STRATEGIES.items():
            variant_corpus_df = make_variant_corpus(base_corpus_df, df, doc_map, args.n_examples,
                                                      enrichment_pool, strat_fn)
            corpus_vecs = embed(variant_corpus_df["text"].tolist(), client, args, cache_dir)

            per_query_df, recall = compute_recall(variant_corpus_df, eval_df, corpus_vecs, query_vecs,
                                                    ks=(1, 10), variant=f"{strat_name}_seed{seed}_doccov")
            per_query_df["strategy"] = strat_name
            per_query_df["seed"] = seed
            covered_per_query.append(per_query_df)

            covered_rows.append({
                "strategy": strat_name, "seed": seed,
                "recall@1": recall[1], "recall@10": recall[10], "n_eval": len(eval_df),
            })

    covered_results_df = pd.DataFrame(covered_rows)
    covered_results_df.to_csv(out_dir / "compare_doc_strategies_doccovered_eval_per_seed.csv", index=False)
    pd.concat(covered_per_query, ignore_index=True).to_csv(
        out_dir / "compare_doc_strategies_doccovered_eval_per_query.csv", index=False)

    print("\n" + "=" * 70)
    print("PASS 2 SUMMARY (mean +/- std across seeds, doc-covered pairs only)")
    print("=" * 70)
    summary2 = covered_results_df.groupby("strategy").agg(
        recall_1_mean=("recall@1", "mean"), recall_1_std=("recall@1", "std"),
        recall_10_mean=("recall@10", "mean"), recall_10_std=("recall@10", "std"),
    ).round(4)
    print(summary2.to_string())
    summary2.to_csv(out_dir / "compare_doc_strategies_doccovered_eval_summary.csv")

    print(f"\nSaved 6 CSVs to {out_dir}/ (per-seed + per-query + summary, x2 passes)")
    print("Read pass-1 summary alongside the reports_recall@1 / non_reports_recall@1 split before")
    print("trusting an overall 'winner' - a strategy can win on average purely by matching")
    print("query_derived's fallback behavior on the 79% Reports-sheet majority.")
    print("Pass-2 summary is the fairer 'is the doc text itself better' comparison, since every")
    print("query there has real (non-fallback) doc text available to doc_description and concat.")


if __name__ == "__main__":
    main()